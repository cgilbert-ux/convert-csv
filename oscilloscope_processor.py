#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Programme de traitement de traces d'oscilloscope
- Sélection de 1 à 4 fichiers CSV (C1, C2, C3, C4)
- Décimation automatique ou manuelle des données
- Export vers Excel avec graphique
- Sauvegarde du paramètre de décimation dans un fichier INI
"""

import os
import sys
import re
import configparser
from pathlib import Path

import pandas as pd
import numpy as np
from tkinter import filedialog, Tk, messagebox, Toplevel, Label, Entry, Button, Frame
from tqdm import tqdm


def get_ini_path():
    """Retourne le chemin du fichier INI (dans le même dossier que le script)"""
    script_dir = Path(__file__).parent.resolve()
    return script_dir / "oscilloscope_config.ini"


def load_decimation_from_ini():
    """Charge la valeur de décimation depuis le fichier INI"""
    ini_path = get_ini_path()
    config = configparser.ConfigParser()
    
    if ini_path.exists():
        config.read(ini_path, encoding='utf-8')
        if 'SETTINGS' in config and 'decimation' in config['SETTINGS']:
            try:
                return int(config['SETTINGS']['decimation'])
            except ValueError:
                pass
    return None


def save_decimation_to_ini(decimation):
    """Sauvegarde la valeur de décimation dans le fichier INI"""
    ini_path = get_ini_path()
    config = configparser.ConfigParser()
    
    if ini_path.exists():
        config.read(ini_path, encoding='utf-8')
    
    if 'SETTINGS' not in config:
        config['SETTINGS'] = {}
    
    config['SETTINGS']['decimation'] = str(decimation)
    
    with open(ini_path, 'w', encoding='utf-8') as f:
        config.write(f)


def ask_decimation_dialog(current_decimation=None):
    """Affiche une boîte de dialogue pour saisir le facteur de décimation"""
    root = Tk()
    root.withdraw()
    root.attributes('-topmost', True)
    
    dialog = Tk()
    dialog.title("Facteur de décimation")
    dialog.attributes('-topmost', True)
    
    label = tk.Label(dialog, text="Entrez le facteur de décimation:")
    label.pack(padx=20, pady=10)
    
    entry = tk.Entry(dialog, width=20)
    if current_decimation is not None:
        entry.insert(0, str(current_decimation))
    entry.pack(padx=20, pady=10)
    
    result = {'value': None}
    
    def on_ok():
        try:
            val = int(entry.get())
            if val >= 1:
                result['value'] = val
                dialog.destroy()
            else:
                messagebox.showerror("Erreur", "La décimation doit être >= 1")
        except ValueError:
            messagebox.showerror("Erreur", "Veuillez entrer un nombre entier valide")
    
    def on_cancel():
        result['value'] = None
        dialog.destroy()
    
    btn_frame = tk.Frame(dialog)
    btn_frame.pack(pady=10)
    
    tk.Button(btn_frame, text="OK", command=on_ok).pack(side=tk.LEFT, padx=5)
    tk.Button(btn_frame, text="Annuler", command=on_cancel).pack(side=tk.LEFT, padx=5)
    
    dialog.bind('<Return>', lambda e: on_ok())
    dialog.bind('<Escape>', lambda e: on_cancel())
    
    dialog.mainloop()
    root.destroy()
    
    return result['value']


def select_csv_files():
    """Affiche une boîte de dialogue pour sélectionner 1 à 4 fichiers CSV"""
    root = Tk()
    root.withdraw()
    root.attributes('-topmost', True)
    
    files = filedialog.askopenfilenames(
        title="Sélectionnez 1 à 4 fichiers CSV (traces d'oscilloscope)",
        filetypes=[("Fichiers CSV", "*.csv"), ("Tous les fichiers", "*.*")],
        multiple=True
    )
    
    root.destroy()
    
    if len(files) == 0:
        return []
    
    if len(files) > 4:
        messagebox.showerror("Erreur", "Vous ne pouvez sélectionner que 1 à 4 fichiers maximum.")
        return []
    
    return list(files)


def extract_channel_info(filepath):
    """Extrait les informations de canal depuis le nom du fichier"""
    filename = os.path.basename(filepath)
    
    # Recherche de motifs comme C1, C2, C3, C4 (case insensitive)
    match = re.search(r'C([1-4])', filename, re.IGNORECASE)
    if match:
        channel_num = int(match.group(1))
        channel_name = f"Voie {channel_num}"
    else:
        channel_num = None
        channel_name = "Trace"
    
    # Nom de base sans information de canal
    base_name = re.sub(r'_?C[1-4]', '', filename, flags=re.IGNORECASE)
    base_name = re.sub(r'\.csv$', '', base_name, flags=re.IGNORECASE)
    
    return {
        'channel_num': channel_num,
        'channel_name': channel_name,
        'base_name': base_name
    }


def calculate_optimal_decimation(total_points, min_lines=50000, max_lines=100000):
    """Calcule la décimation optimale pour avoir entre min_lines et max_lines"""
    if total_points <= max_lines:
        return 1
    
    # On vise environ 75000 lignes (milieu de la plage)
    target_lines = (min_lines + max_lines) // 2
    decimation = max(1, total_points // target_lines)
    
    # Vérifier que le résultat est dans la plage acceptable
    resulting_lines = total_points // decimation
    
    if resulting_lines < min_lines and decimation > 1:
        decimation -= 1
    
    return max(1, decimation)


def read_csv_file(filepath):
    """Lit un fichier CSV d'oscilloscope et retourne un DataFrame"""
    # Essayer différents encodages et séparateurs
    encodings = ['utf-8', 'latin-1', 'cp1252']
    separators = [',', ';', '\t']
    
    for encoding in encodings:
        for sep in separators:
            try:
                df = pd.read_csv(filepath, encoding=encoding, sep=sep, 
                               skiprows=0, low_memory=False)
                if len(df.columns) >= 2:
                    return df, sep, encoding
            except Exception:
                continue
    
    # Si échec, essayer sans spécifier l'encodage
    try:
        df = pd.read_csv(filepath, low_memory=False, engine='python')
        return df, ',', 'utf-8'
    except Exception as e:
        raise Exception(f"Impossible de lire le fichier {filepath}: {str(e)}")


def decimate_data(df, decimation_factor):
    """Réduit le nombre de points par décimation"""
    if decimation_factor <= 1:
        return df
    
    # Prendre un point sur N
    decimated_df = df.iloc[::decimation_factor].reset_index(drop=True)
    return decimated_df


def create_excel_with_chart(data_dict, output_path, progress_callback=None):
    """Crée un fichier Excel avec les données et un graphique"""
    from openpyxl import Workbook
    from openpyxl.chart import LineChart, Reference
    from openpyxl.styles import Font, Alignment
    from openpyxl.utils.dataframe import dataframe_to_rows
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Données"
    
    # En-têtes
    channels = list(data_dict.keys())
    headers = ["Point"] + [f"{ch['channel_name']} (V)" for ch in channels]
    
    if data_dict:
        first_channel = channels[0]
        time_data = data_dict[first_channel]['time']
    else:
        time_data = []
    
    # Écriture des en-têtes
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
    
    total_rows = len(time_data) if time_data is not None else 0
    
    # Écriture des données avec barre de progression
    for row_idx in tqdm(range(total_rows), desc="Écriture des données Excel", 
                       ncols=100, bar_format='{l_bar}{bar}| {n_fmt}/{total_fmt}'):
        ws.cell(row=row_idx + 2, column=1, value=time_data[row_idx] if row_idx < len(time_data) else None)
        
        for col_idx, channel in enumerate(channels, 2):
            voltage_data = data_dict[channel]['voltage']
            value = voltage_data[row_idx] if row_idx < len(voltage_data) else None
            ws.cell(row=row_idx + 2, column=col_idx, value=value)
        
        if progress_callback and row_idx % 1000 == 0:
            progress_callback(row_idx, total_rows)
    
    # Création du graphique
    chart = LineChart()
    chart.title = "Traces d'oscilloscope"
    chart.style = 13
    chart.y_axis.title = 'Tension (V)'
    chart.x_axis.title = 'Point de mesure'
    chart.width = 20
    chart.height = 10
    
    # Ajout des séries pour chaque voie
    for col_idx, channel in enumerate(channels, 2):
        values = Reference(ws, min_col=col_idx, min_row=2, max_row=total_rows + 1)
        series_name = ws.cell(row=1, column=col_idx).value
        series = Series(values, from_series=series_name)
        chart.series.append(series)
    
    ws.add_chart(chart, "E2")
    
    wb.save(output_path)
    return True


class SimpleProgressBar:
    """Barre de progression simple pour console"""
    def __init__(self, total, desc="Progression"):
        self.total = total
        self.desc = desc
        self.current = 0
    
    def update(self, current, total=None):
        if total:
            self.total = total
        self.current = current
        percent = (self.current / self.total * 100) if self.total > 0 else 0
        bar_length = 40
        filled = int(bar_length * self.current / self.total) if self.total > 0 else 0
        bar = '█' * filled + '-' * (bar_length - filled)
        print(f"\r{self.desc}: [{bar}] {percent:.1f}% ({current}/{total})", end='', flush=True)
        if self.current >= self.total:
            print()


def main():
    """Fonction principale"""
    import tkinter as tk
    
    print("=" * 60)
    print("Traitement de traces d'oscilloscope")
    print("=" * 60)
    
    # Étape 1: Sélection des fichiers
    print("\n[Sélection des fichiers]")
    csv_files = select_csv_files()
    
    if not csv_files:
        print("Aucun fichier sélectionné. Abandon.")
        return
    
    print(f"{len(csv_files)} fichier(s) sélectionné(s):")
    for f in csv_files:
        print(f"  - {os.path.basename(f)}")
    
    # Étape 2: Charger la décimation depuis INI
    saved_decimation = load_decimation_from_ini()
    
    # Étape 3: Demander confirmation/modification de la décimation
    print("\n[Configuration de la décimation]")
    
    # D'abord, lire un fichier pour connaître le nombre de points
    try:
        sample_df, _, _ = read_csv_file(csv_files[0])
        total_points = len(sample_df)
        print(f"Nombre de points dans le premier fichier: {total_points:,}")
        
        # Calculer la décimation automatique
        auto_decimation = calculate_optimal_decimation(total_points)
        print(f"Décimation automatique recommandée: {auto_decimation}")
        print(f"  -> Nombre de points résultant: {total_points // auto_decimation:,}")
        
        # Utiliser la valeur sauvegardée ou la valeur automatique
        initial_decimation = saved_decimation if saved_decimation else auto_decimation
        
    except Exception as e:
        print(f"Erreur lors de la lecture du fichier: {e}")
        initial_decimation = saved_decimation if saved_decimation else 100
    
    # Boîte de dialogue pour la décimation
    root = Tk()
    root.withdraw()
    root.attributes('-topmost', True)
    
    dialog = Toplevel(root)
    dialog.title("Facteur de décimation")
    dialog.attributes('-topmost', True)
    
    label = tk.Label(dialog, text=f"Entrez le facteur de décimation:\n(Recommandé: {initial_decimation})")
    label.pack(padx=20, pady=10)
    
    entry = tk.Entry(dialog, width=20)
    entry.insert(0, str(initial_decimation))
    entry.pack(padx=20, pady=10)
    
    result = {'value': None}
    
    def on_ok():
        try:
            val = int(entry.get())
            if val >= 1:
                result['value'] = val
                dialog.destroy()
            else:
                messagebox.showerror("Erreur", "La décimation doit être >= 1")
        except ValueError:
            messagebox.showerror("Erreur", "Veuillez entrer un nombre entier valide")
    
    def on_auto():
        entry.delete(0, tk.END)
        entry.insert(0, str(initial_decimation))
    
    def on_cancel():
        result['value'] = initial_decimation
        dialog.destroy()
    
    btn_frame = tk.Frame(dialog)
    btn_frame.pack(pady=10)
    
    tk.Button(btn_frame, text="Automatique", command=on_auto).pack(side=tk.LEFT, padx=5)
    tk.Button(btn_frame, text="OK", command=on_ok).pack(side=tk.LEFT, padx=5)
    tk.Button(btn_frame, text="Annuler", command=on_cancel).pack(side=tk.LEFT, padx=5)
    
    dialog.bind('<Return>', lambda e: on_ok())
    dialog.bind('<Escape>', lambda e: on_cancel())
    
    dialog.mainloop()
    root.destroy()
    
    decimation_factor = result['value'] if result['value'] else initial_decimation
    
    # Sauvegarder dans INI
    save_decimation_to_ini(decimation_factor)
    print(f"Décimation utilisée: {decimation_factor}")
    print(f"Paramètre sauvegardé dans: {get_ini_path()}")
    
    # Étape 4: Lecture et traitement des fichiers
    print("\n[Lecture et décimation des fichiers]")
    
    channel_data = {}
    
    for filepath in tqdm(csv_files, desc="Lecture des fichiers", ncols=100):
        try:
            df, sep, encoding = read_csv_file(filepath)
            channel_info = extract_channel_info(filepath)
            
            # Identifier les colonnes (temps et tension)
            if len(df.columns) >= 2:
                time_col = df.columns[0]
                voltage_col = df.columns[1]
                
                # Décimation
                df_decimated = decimate_data(df, decimation_factor)
                
                channel_key = f"C{channel_info['channel_num']}" if channel_info['channel_num'] else filepath
                
                channel_data[channel_key] = {
                    'time': df_decimated[time_col].values,
                    'voltage': df_decimated[voltage_col].values,
                    'channel_info': channel_info,
                    'original_file': filepath
                }
                
                print(f"  {os.path.basename(filepath)}: {len(df)} -> {len(df_decimated)} points")
                
        except Exception as e:
            print(f"Erreur lors du traitement de {filepath}: {e}")
            continue
    
    if not channel_data:
        print("Aucune donnée traitée avec succès. Abandon.")
        return
    
    # Étape 5: Génération du nom de fichier Excel
    first_channel = list(channel_data.values())[0]
    base_name = first_channel['channel_info']['base_name']
    
    # Nettoyer le nom de base
    base_name = re.sub(r'[^\w\-_]', '_', base_name)
    base_name = re.sub(r'_+', '_', base_name)
    base_name = base_name.strip('_')
    
    if not base_name:
        base_name = "oscilloscope_data"
    
    output_filename = f"{base_name}_traces.xlsx"
    output_path = os.path.join(os.path.dirname(csv_files[0]), output_filename)
    
    print(f"\n[Création du fichier Excel]")
    print(f"Nom du fichier: {output_filename}")
    
    # Étape 6: Création du fichier Excel avec graphique
    from openpyxl import Workbook
    from openpyxl.chart import LineChart, Reference, Series
    from openpyxl.styles import Font, Alignment
    from openpyxl.utils.dataframe import dataframe_to_rows
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Données"
    
    # Préparer les données
    channels_list = list(channel_data.keys())
    first_channel_data = channel_data[channels_list[0]]
    time_data = first_channel_data['time']
    
    # En-têtes
    headers = ["Point"]
    for ch_key in channels_list:
        ch_info = channel_data[ch_key]['channel_info']
        headers.append(f"{ch_info['channel_name']} (V)")
    
    print("Écriture des en-têtes...")
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
    
    # Écriture des données avec barre de progression
    total_rows = len(time_data)
    print(f"Écriture de {total_rows:,} lignes de données...")
    
    progress_bar = SimpleProgressBar(total_rows, "Progression")
    
    for row_idx in range(total_rows):
        ws.cell(row=row_idx + 2, column=1, value=float(time_data[row_idx]))
        
        for col_idx, ch_key in enumerate(channels_list, 2):
            voltage_data = channel_data[ch_key]['voltage']
            if row_idx < len(voltage_data):
                try:
                    value = float(voltage_data[row_idx])
                except (ValueError, TypeError):
                    value = None
            else:
                value = None
            ws.cell(row=row_idx + 2, column=col_idx, value=value)
        
        if row_idx % 1000 == 0:
            progress_bar.update(row_idx, total_rows)
    
    progress_bar.update(total_rows, total_rows)
    
    # Création du graphique
    print("Création du graphique...")
    chart = LineChart()
    chart.title = "Traces d'oscilloscope"
    chart.style = 13
    chart.y_axis.title = 'Tension (V)'
    chart.x_axis.title = 'Point de mesure'
    chart.width = 20
    chart.height = 10
    
    # Ajout des séries pour chaque voie
    for col_idx, ch_key in enumerate(channels_list, 2):
        ch_info = channel_data[ch_key]['channel_info']
        values = Reference(ws, min_col=col_idx, min_row=2, max_row=total_rows + 1)
        series = Series(values, title=ch_info['channel_name'])
        chart.series.append(series)
    
    ws.add_chart(chart, "E2")
    
    # Sauvegarde
    print(f"Sauvegarde dans {output_path}...")
    wb.save(output_path)
    
    print("\n" + "=" * 60)
    print("TRAITEMENT TERMINÉ AVEC SUCCÈS")
    print("=" * 60)
    print(f"Fichier Excel créé: {output_path}")
    print(f"Nombre de traces: {len(channels_list)}")
    print(f"Nombre de points après décimation: {total_rows:,}")
    print(f"Facteur de décimation: {decimation_factor}")
    print(f"Configurations sauvegardées dans: {get_ini_path()}")
    print("=" * 60)


if __name__ == "__main__":
    main()
